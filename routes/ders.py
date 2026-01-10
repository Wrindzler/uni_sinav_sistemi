"""
DERS YÖNETİMİ ROUTE'LARI
========================
Ders ekleme, silme, güncelleme işlemleri.

Bölüm yetkilileri ve admin kullanıcılar ders yönetimi yapabilir.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from models.database import db
from models.ders import Ders
from models.bolum import Bolum
from models.ogretim_uyesi import OgretimUyesi
from utils.decorators import bolum_yetkilisi_required
from utils.helpers import get_bolum_yetkilisi_bolum_id

# Blueprint oluştur
ders_bp = Blueprint('ders', __name__, url_prefix='/ders')


def _get_bolumler_ogretim_uyeleri(current_user):
    """
    Kullanıcının yetkisine göre bölümler ve öğretim üyelerini getir.
    
    Returns:
        tuple: (bolumler, ogretim_uyeleri)
    """
    if current_user.is_admin():
        bolumler = Bolum.query.all()
        ogretim_uyeleri = OgretimUyesi.query.all()
    else:
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if bolum_id:
            bolumler = Bolum.query.filter_by(id=bolum_id).all()
            ogretim_uyeleri = OgretimUyesi.query.filter_by(bolum_id=bolum_id).all()
        else:
            bolumler = []
            ogretim_uyeleri = []
    
    return bolumler, ogretim_uyeleri


@ders_bp.route('/')
@login_required
@bolum_yetkilisi_required
def liste():
    """
    Ders listesi sayfası.
    
    Kullanıcının yetkili olduğu bölümlere ait dersleri gösterir.
    """
    # Admin ise tüm dersler, bölüm yetkilisi ise sadece kendi bölümü
    if current_user.is_admin():
        dersler = Ders.query.all()
    else:
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if bolum_id:
            dersler = Ders.query.filter_by(bolum_id=bolum_id, aktif=True).all()
        else:
            dersler = []
    
    return render_template('ders/liste.html', dersler=dersler)


@ders_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
@bolum_yetkilisi_required
def ekle():
    """
    Yeni ders ekleme sayfası.
    
    GET: Ders ekleme formunu gösterir
    POST: Yeni ders oluşturur
    """
    if request.method == 'POST':
        # Form verilerini al
        ad = request.form.get('ad')
        kod = request.form.get('kod')

        try:
            bolum_id = int(request.form.get('bolum_id'))
            ogretim_uyesi_id = int(request.form.get('ogretim_uyesi_id'))
            ogrenci_sayisi = int(request.form.get('ogrenci_sayisi', 0))
            sinav_suresi = int(request.form.get('sinav_suresi', 60))

            # Validate positive values
            if ogrenci_sayisi < 0:
                flash('Öğrenci sayısı negatif olamaz!', 'danger')
                raise ValueError('Invalid student count')
            if sinav_suresi <= 0:
                flash('Sınav süresi pozitif bir sayı olmalıdır!', 'danger')
                raise ValueError('Invalid exam duration')
        except (ValueError, TypeError):
            # Re-render form with data
            bolumler, ogretim_uyeleri = _get_bolumler_ogretim_uyeleri(current_user)
            return render_template('ders/ekle.html', bolumler=bolumler, ogretim_uyeleri=ogretim_uyeleri)

        sinav_turu = request.form.get('sinav_turu', 'yazili')

        # Yeni ders oluştur
        ders = Ders(
            ad=ad,
            kod=kod,
            bolum_id=bolum_id,
            ogretim_uyesi_id=ogretim_uyesi_id,
            ogrenci_sayisi=ogrenci_sayisi,
            sinav_suresi=sinav_suresi,
            sinav_turu=sinav_turu
        )

        db.session.add(ders)
        db.session.commit()

        flash('Ders başarıyla eklendi!', 'success')
        return redirect(url_for('ders.liste'))

    # Form için gerekli veriler
    bolumler, ogretim_uyeleri = _get_bolumler_ogretim_uyeleri(current_user)
    return render_template('ders/ekle.html', bolumler=bolumler, ogretim_uyeleri=ogretim_uyeleri)


@ders_bp.route('/<int:ders_id>/duzenle', methods=['GET', 'POST'])
@login_required
@bolum_yetkilisi_required
def duzenle(ders_id):
    """
    Ders düzenleme sayfası.
    
    Args:
        ders_id: Düzenlenecek ders ID'si
    """
    ders = Ders.query.get_or_404(ders_id)
    
    # Yetki kontrolü
    if not current_user.is_admin():
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if not bolum_id or ders.bolum_id != bolum_id:
            flash('Bu dersi düzenleme yetkiniz yok!', 'error')
            return redirect(url_for('ders.liste'))
    
    if request.method == 'POST':
        # Form verilerini güncelle
        ders.ad = request.form.get('ad')
        ders.kod = request.form.get('kod')

        try:
            bolum_id = int(request.form.get('bolum_id'))
            ogretim_uyesi_id = int(request.form.get('ogretim_uyesi_id'))
            ogrenci_sayisi = int(request.form.get('ogrenci_sayisi', 0))
            sinav_suresi = int(request.form.get('sinav_suresi', 60))

            # Validate positive values
            if ogrenci_sayisi < 0:
                flash('Öğrenci sayısı negatif olamaz!', 'danger')
                raise ValueError('Invalid student count')
            if sinav_suresi <= 0:
                flash('Sınav süresi pozitif bir sayı olmalıdır!', 'danger')
                raise ValueError('Invalid exam duration')

            ders.bolum_id = bolum_id
            ders.ogretim_uyesi_id = ogretim_uyesi_id
            ders.ogrenci_sayisi = ogrenci_sayisi
            ders.sinav_suresi = sinav_suresi
        except (ValueError, TypeError):
            # Re-render form with existing data
            bolumler, ogretim_uyeleri = _get_bolumler_ogretim_uyeleri(current_user)
            return render_template('ders/duzenle.html', ders=ders, bolumler=bolumler, ogretim_uyeleri=ogretim_uyeleri)

        ders.sinav_turu = request.form.get('sinav_turu', 'yazili')

        db.session.commit()

        flash('Ders başarıyla güncellendi!', 'success')
        return redirect(url_for('ders.liste'))
    
    # Form için gerekli veriler
    bolumler, ogretim_uyeleri = _get_bolumler_ogretim_uyeleri(current_user)
    return render_template('ders/duzenle.html', ders=ders, bolumler=bolumler, ogretim_uyeleri=ogretim_uyeleri)


@ders_bp.route('/<int:ders_id>/sil', methods=['POST'])
@login_required
@bolum_yetkilisi_required
def sil(ders_id):
    """
    Ders silme işlemi.
    
    Args:
        ders_id: Silinecek ders ID'si
    """
    from models.ogrenci_ders import OgrenciDers
    from models.sinav import Sinav
    from models.ozel_durum import OzelDurum
    
    ders = Ders.query.get_or_404(ders_id)
    
    # Yetki kontrolü
    if not current_user.is_admin():
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if not bolum_id or ders.bolum_id != bolum_id:
            flash('Bu dersi silme yetkiniz yok!', 'error')
            return redirect(url_for('ders.liste'))
    
    try:
        ders_adi = ders.ad
        
        # İlişkili kayıtları sil
        OgrenciDers.query.filter_by(ders_id=ders_id).delete()
        Sinav.query.filter_by(ders_id=ders_id).delete()
        OzelDurum.query.filter_by(ders_id=ders_id).delete()
        
        # Dersi sil
        db.session.delete(ders)
        db.session.commit()
        
        flash(f'"{ders_adi}" dersi başarıyla silindi!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ders silinirken hata oluştu: {str(e)}', 'error')
    
    return redirect(url_for('ders.liste'))


@ders_bp.route('/<int:ders_id>/ogrenci-yukle', methods=['GET', 'POST'])
@login_required
@bolum_yetkilisi_required
def ogrenci_yukle(ders_id):
    """
    Derse özel öğrenci listesi yükleme.
    
    CSV veya Excel dosyasından öğrencileri yükler ve bu derse kaydeder.
    
    Args:
        ders_id: Öğrenci eklenecek ders ID'si
    """
    from models.ogrenci import Ogrenci
    from models.ogrenci_ders import OgrenciDers
    from werkzeug.utils import secure_filename
    import os
    import csv
    from openpyxl import load_workbook
    from config import Config
    
    ders = Ders.query.get_or_404(ders_id)
    
    # Yetki kontrolü
    if not current_user.is_admin():
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if not bolum_id or ders.bolum_id != bolum_id:
            flash('Bu derse öğrenci yükleme yetkiniz yok!', 'error')
            return redirect(url_for('ders.liste'))
    
    if request.method == 'POST':
        # Dosya kontrolü
        if 'dosya' not in request.files:
            flash('Dosya seçilmedi!', 'error')
            return redirect(url_for('ders.ogrenci_yukle', ders_id=ders_id))
        
        dosya = request.files['dosya']
        
        if dosya.filename == '':
            flash('Dosya seçilmedi!', 'error')
            return redirect(url_for('ders.ogrenci_yukle', ders_id=ders_id))
        
        # Dosya uzantısı kontrolü
        if not ('.' in dosya.filename and dosya.filename.rsplit('.', 1)[1].lower() in ['csv', 'xlsx', 'xls']):
            flash('Geçersiz dosya formatı! Sadece CSV, XLSX veya XLS dosyaları yüklenebilir.', 'error')
            return redirect(url_for('ders.ogrenci_yukle', ders_id=ders_id))
        
        # Uploads klasörünü oluştur (yoksa)
        uploads_dir = Config.UPLOAD_FOLDER
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)
        
        # Dosyayı kaydet
        filename = secure_filename(dosya.filename)
        dosya_yolu = os.path.join(uploads_dir, filename)
        dosya.save(dosya_yolu)
        
        # Dosyayı işle
        try:
            sonuc = _dosyadan_ogrenci_yukle_ve_derse_kaydet(dosya_yolu, filename, ders)
            
            # Dosyayı sil (hata olsa bile devam et)
            try:
                os.remove(dosya_yolu)
            except PermissionError:
                # Windows'ta dosya hala açıksa silme, sorun değil
                pass
            
            # Dersin öğrenci sayısını güncelle
            ders.ogrenci_sayisi = OgrenciDers.query.filter_by(ders_id=ders.id).count()
            db.session.commit()
            
            flash(f'{sonuc["eklenen"]} öğrenci derse kaydedildi, {sonuc["hata"]} hata oluştu.', 
                  'success' if sonuc['hata'] == 0 else 'warning')
            return redirect(url_for('ders.liste'))
        except Exception as e:
            # Dosyayı sil (hata olsa bile devam et)
            try:
                if os.path.exists(dosya_yolu):
                    os.remove(dosya_yolu)
            except PermissionError:
                pass
            flash(f'Dosya işlenirken hata oluştu: {str(e)}', 'error')
            return redirect(url_for('ders.ogrenci_yukle', ders_id=ders_id))
    
    # GET isteği: Yükleme formunu göster
    # Dersin mevcut öğrenci sayısını göster
    mevcut_ogrenci_sayisi = OgrenciDers.query.filter_by(ders_id=ders.id).count()
    
    return render_template('ders/ogrenci_yukle.html', 
                         ders=ders,
                         mevcut_ogrenci_sayisi=mevcut_ogrenci_sayisi)


@ders_bp.route('/toplu-yukle', methods=['GET', 'POST'])
@login_required
@bolum_yetkilisi_required
def toplu_yukle():
    """
    Excel dosyasından ders ve öğrencileri toplu yükleme.
    
    Dosya adından ders kodunu çıkarır (örn: SınıfListesi[BLM111].xls -> BLM111)
    Ders yoksa oluşturur, öğrencileri ekler ve derse kaydeder.
    """
    from models.ogrenci import Ogrenci
    from models.ogrenci_ders import OgrenciDers
    from werkzeug.utils import secure_filename
    import os
    import re
    from config import Config
    
    if request.method == 'POST':
        # Dosya kontrolü
        if 'dosya' not in request.files:
            flash('Dosya seçilmedi!', 'error')
            return redirect(url_for('ders.toplu_yukle'))
        
        dosya = request.files['dosya']
        
        if dosya.filename == '':
            flash('Dosya seçilmedi!', 'error')
            return redirect(url_for('ders.toplu_yukle'))
        
        # Dosya uzantısı kontrolü
        if not ('.' in dosya.filename and dosya.filename.rsplit('.', 1)[1].lower() in ['xlsx', 'xls']):
            flash('Geçersiz dosya formatı! Sadece XLSX veya XLS dosyaları yüklenebilir.', 'error')
            return redirect(url_for('ders.toplu_yukle'))
        
        # Dosya adından ders kodunu çıkar (örn: SınıfListesi[BLM111].xls -> BLM111)
        original_filename = dosya.filename
        ders_kodu_match = re.search(r'\[([^\]]+)\]', original_filename)
        if not ders_kodu_match:
            flash('Dosya adında ders kodu bulunamadı! Dosya adı SınıfListesi[DERSKODU].xls formatında olmalı.', 'error')
            return redirect(url_for('ders.toplu_yukle'))
        
        ders_kodu = ders_kodu_match.group(1).strip()
        
        # Uploads klasörünü oluştur (yoksa)
        uploads_dir = Config.UPLOAD_FOLDER
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)
        
        # Dosyayı kaydet
        filename = secure_filename(dosya.filename)
        dosya_yolu = os.path.join(uploads_dir, filename)
        dosya.save(dosya_yolu)
        
        # Dosyayı işle
        try:
            sonuc = _excel_den_ders_ve_ogrenci_yukle(dosya_yolu, filename, ders_kodu, current_user)
            
            # Dosyayı sil
            try:
                os.remove(dosya_yolu)
            except PermissionError:
                pass
            
            flash(f'Ders "{sonuc["ders_adi"]}" oluşturuldu/güncellendi. {sonuc["eklenen"]} öğrenci eklendi, {sonuc["mevcut"]} öğrenci zaten kayıtlıydı.', 'success')
            return redirect(url_for('ders.liste'))
        except Exception as e:
            try:
                if os.path.exists(dosya_yolu):
                    os.remove(dosya_yolu)
            except PermissionError:
                pass
            flash(f'Dosya işlenirken hata oluştu: {str(e)}', 'error')
            return redirect(url_for('ders.toplu_yukle'))
    
    # GET: Bölümleri getir
    if current_user.is_admin():
        bolumler = Bolum.query.all()
    else:
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if bolum_id:
            bolumler = Bolum.query.filter_by(id=bolum_id).all()
        else:
            bolumler = []
    
    return render_template('ders/toplu_yukle.html', bolumler=bolumler)


def _excel_den_ders_ve_ogrenci_yukle(dosya_yolu, filename, ders_kodu, current_user):
    """
    Excel dosyasından ders ve öğrencileri yükle.
    
    Args:
        dosya_yolu: Excel dosyasının yolu
        filename: Dosya adı
        ders_kodu: Ders kodu (dosya adından çıkarılmış)
        current_user: Mevcut kullanıcı
        
    Returns:
        dict: {'ders_adi': str, 'eklenen': int, 'mevcut': int, 'hata': int}
    """
    from models.ogrenci import Ogrenci
    from models.ogrenci_ders import OgrenciDers
    import xlrd
    
    eklenen = 0
    mevcut = 0
    hata = 0
    bolum_adi = None
    
    # Excel dosyasını oku
    wb = xlrd.open_workbook(dosya_yolu)
    ws = wb.sheet_by_index(0)
    
    # Başlık satırını ve sütun indekslerini bul
    baslik_satiri = -1
    ogrenci_no_col = -1
    ad_soyad_col = -1
    bolum_col = -1
    
    for i in range(min(20, ws.nrows)):
        for j in range(ws.ncols):
            cell_val = str(ws.cell_value(i, j)).lower().strip()
            if 'renci no' in cell_val or cell_val == '#':
                if cell_val == '#':
                    baslik_satiri = i
                else:
                    baslik_satiri = i
                    ogrenci_no_col = j
            if 'soyad' in cell_val:
                ad_soyad_col = j
            if 'l' in cell_val and 'm' in cell_val and len(cell_val) < 10:  # Bölüm sütunu
                bolum_col = j
        if baslik_satiri >= 0 and ogrenci_no_col < 0:
            ogrenci_no_col = 4
            ad_soyad_col = 5
            bolum_col = 1
        if baslik_satiri >= 0:
            break
    
    if baslik_satiri < 0:
        baslik_satiri = 4
        ogrenci_no_col = 4
        ad_soyad_col = 5
        bolum_col = 1
    
    # İlk veri satırından bölüm adını al
    baslangic = baslik_satiri + 1
    if baslangic < ws.nrows:
        bolum_adi = str(ws.cell_value(baslangic, bolum_col)).strip() if ws.cell_value(baslangic, bolum_col) else None
    
    # Bölümü ders koduna göre belirle
    bolum_id = None
    
    # Ders koduna göre otomatik bölüm ataması
    if ders_kodu:
        ders_kodu_upper = ders_kodu.upper()
        if ders_kodu_upper.startswith('YZM'):
            # YZM ile başlıyorsa Yazılım Mühendisliği
            yazilim_bolum = Bolum.query.filter(Bolum.ad.ilike('%Yazılım%Mühendisliği%')).first()
            if yazilim_bolum:
                bolum_id = yazilim_bolum.id
        elif ders_kodu_upper.startswith('BLM'):
            # BLM ile başlıyorsa Bilgisayar Mühendisliği
            bilgisayar_bolum = Bolum.query.filter(Bolum.ad.ilike('%Bilgisayar%Mühendisliği%')).first()
            if bilgisayar_bolum:
                bolum_id = bilgisayar_bolum.id
        elif ders_kodu_upper.startswith('MAT') or ders_kodu_upper.startswith('FIZ') or ders_kodu_upper.startswith('SEC'):
            # Ortak dersler (MAT, FIZ, SEC) - Bilgisayar Mühendisliğine ata (varsayılan)
            bilgisayar_bolum = Bolum.query.filter(Bolum.ad.ilike('%Bilgisayar%Mühendisliği%')).first()
            if bilgisayar_bolum:
                bolum_id = bilgisayar_bolum.id
    
    # Ders kodundan bölüm bulunamazsa, eski yöntemi dene
    if not bolum_id:
        if current_user.is_admin():
            # Admin için bölümü bulmaya çalış veya ilk bölümü kullan
            if bolum_adi:
                bolum = Bolum.query.filter(Bolum.ad.ilike(f'%{bolum_adi[:20]}%')).first()
                if bolum:
                    bolum_id = bolum.id
            if not bolum_id:
                bolum = Bolum.query.first()
                if bolum:
                    bolum_id = bolum.id
        else:
            bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
    
    if not bolum_id:
        raise Exception('Bölüm bulunamadı!')
    
    # Bölüme ait öğretim üyesini bul, yoksa herhangi birini al
    ogretim_uyesi = OgretimUyesi.query.filter_by(bolum_id=bolum_id).first()
    if not ogretim_uyesi:
        ogretim_uyesi = OgretimUyesi.query.first()
    
    if not ogretim_uyesi:
        raise Exception('Öğretim üyesi bulunamadı! Önce bir öğretim üyesi ekleyin.')
    
    # Dersi bul veya oluştur
    ders = Ders.query.filter_by(kod=ders_kodu).first()
    if not ders:
        ders = Ders(
            kod=ders_kodu,
            ad=ders_kodu,  # Başlangıçta kod ile aynı, sonra düzenlenebilir
            bolum_id=bolum_id,
            ogretim_uyesi_id=ogretim_uyesi.id,
            ogrenci_sayisi=0,
            sinav_suresi=60,
            sinav_turu='yazili',
            aktif=True
        )
        db.session.add(ders)
        db.session.flush()
    
    # Öğrencileri oku ve ekle
    for satir in range(baslangic, ws.nrows):
        # Öğrenci numarasını al
        ogrenci_no_raw = ws.cell_value(satir, ogrenci_no_col)
        if isinstance(ogrenci_no_raw, float):
            ogrenci_no = str(int(ogrenci_no_raw)).strip()
        else:
            ogrenci_no = str(ogrenci_no_raw).strip() if ogrenci_no_raw else None
        
        # Boş veya geçersiz öğrenci numarası atla
        if not ogrenci_no or ogrenci_no == '0' or ogrenci_no == '':
            continue
        
        # Ad Soyadı al ve ayır
        ad_soyad = str(ws.cell_value(satir, ad_soyad_col)).strip() if ws.cell_value(satir, ad_soyad_col) else None
        
        if not ad_soyad:
            hata += 1
            continue
        
        # Ad Soyadı ayır (son kelime soyad, kalanlar ad)
        parcalar = ad_soyad.split()
        if len(parcalar) >= 2:
            soyad = parcalar[-1]
            ad = ' '.join(parcalar[:-1])
        else:
            ad = ad_soyad
            soyad = ''
        
        if not ad:
            hata += 1
            continue
        
        # Öğrenci var mı kontrol et
        ogrenci = Ogrenci.query.filter_by(ogrenci_no=ogrenci_no).first()
        
        if not ogrenci:
            # Yeni öğrenci oluştur
            ogrenci = Ogrenci(
                ogrenci_no=ogrenci_no,
                ad=ad,
                soyad=soyad,
                email=None,
                bolum_id=bolum_id
            )
            db.session.add(ogrenci)
            db.session.flush()
        
        # Öğrenciyi derse kaydet (zaten kayıtlı değilse)
        if not OgrenciDers.query.filter_by(ogrenci_id=ogrenci.id, ders_id=ders.id).first():
            ogrenci_ders = OgrenciDers(
                ogrenci_id=ogrenci.id,
                ders_id=ders.id
            )
            db.session.add(ogrenci_ders)
            eklenen += 1
        else:
            mevcut += 1
    
    # Dersin öğrenci sayısını güncelle
    ders.ogrenci_sayisi = OgrenciDers.query.filter_by(ders_id=ders.id).count()
    
    db.session.commit()
    
    return {
        'ders_adi': ders.kod,
        'eklenen': eklenen,
        'mevcut': mevcut,
        'hata': hata
    }


def _dosyadan_ogrenci_yukle_ve_derse_kaydet(dosya_yolu, filename, ders):
    """
    Excel veya CSV dosyasından öğrencileri yükle ve derse kaydet.
    
    Dosya formatı:
    - CSV: ogrenci_no,ad,soyad,email (başlık satırı opsiyonel)
    - Excel: A=ogrenci_no, B=ad, C=soyad, D=email (ilk satır başlık olabilir)
    
    Args:
        dosya_yolu: Dosyanın tam yolu
        filename: Dosya adı
        ders: Ders nesnesi
        
    Returns:
        dict: {'eklenen': int, 'hata': int}
    """
    from models.ogrenci import Ogrenci
    from models.ogrenci_ders import OgrenciDers
    import csv
    from openpyxl import load_workbook
    
    eklenen = 0
    hata = 0
    
    # Dosya uzantısına göre işle
    if filename.endswith('.csv'):
        # CSV dosyasını oku
        with open(dosya_yolu, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            satirlar = list(reader)
            
            # İlk satır başlık olabilir, kontrol et
            baslangic = 0
            if satirlar and (satirlar[0][0].lower() in ['ogrenci_no', 'öğrenci_no', 'numara', 'no']):
                baslangic = 1
            
            for satir in satirlar[baslangic:]:
                if len(satir) < 3:  # En az ogrenci_no, ad, soyad olmalı
                    hata += 1
                    continue
                
                ogrenci_no = satir[0].strip()
                ad = satir[1].strip()
                soyad = satir[2].strip()
                email = satir[3].strip() if len(satir) > 3 else None
                
                if not ogrenci_no or not ad or not soyad:
                    hata += 1
                    continue
                
                # Öğrenci var mı kontrol et
                ogrenci = Ogrenci.query.filter_by(ogrenci_no=ogrenci_no).first()
                
                if not ogrenci:
                    # Yeni öğrenci oluştur
                    ogrenci = Ogrenci(
                        ogrenci_no=ogrenci_no,
                        ad=ad,
                        soyad=soyad,
                        email=email,
                        bolum_id=ders.bolum_id
                    )
                    db.session.add(ogrenci)
                    db.session.flush()  # ID'yi al
                
                # Öğrenciyi derse kaydet (zaten kayıtlı değilse)
                if not OgrenciDers.query.filter_by(ogrenci_id=ogrenci.id, ders_id=ders.id).first():
                    ogrenci_ders = OgrenciDers(
                        ogrenci_id=ogrenci.id,
                        ders_id=ders.id
                    )
                    db.session.add(ogrenci_ders)
                    eklenen += 1
    
    elif filename.endswith('.xls'):
        # Eski .xls formatı - xlrd kullan
        import xlrd
        wb = xlrd.open_workbook(dosya_yolu)
        try:
            ws = wb.sheet_by_index(0)
            
            # Başlık satırını ve sütun indekslerini bul
            baslik_satiri = -1
            ogrenci_no_col = -1
            ad_soyad_col = -1
            
            for i in range(min(20, ws.nrows)):  # İlk 20 satırda başlık ara
                for j in range(ws.ncols):
                    cell_val = str(ws.cell_value(i, j)).lower().strip()
                    # Encoding sorunları için farklı varyasyonları dene
                    if 'renci no' in cell_val or cell_val == '#':
                        if cell_val == '#':
                            # # sütunu bulundu, öğrenci no genelde 4. sütunda
                            baslik_satiri = i
                        else:
                            baslik_satiri = i
                            ogrenci_no_col = j
                    if 'soyad' in cell_val:
                        ad_soyad_col = j
                if baslik_satiri >= 0 and ogrenci_no_col < 0:
                    # # sütunu bulundu ama öğrenci no sütunu bulunamadı, varsayılan kullan
                    ogrenci_no_col = 4  # Genelde E sütunu
                    ad_soyad_col = 5    # Genelde F sütunu
                if baslik_satiri >= 0:
                    break
            
            # Başlık bulunamazsa varsayılan değerler kullan
            if baslik_satiri < 0:
                baslik_satiri = 4  # Genelde 5. satır başlık
                ogrenci_no_col = 4
                ad_soyad_col = 5
            
            baslangic = baslik_satiri + 1
            
            for satir in range(baslangic, ws.nrows):
                # Öğrenci numarasını al
                ogrenci_no_raw = ws.cell_value(satir, ogrenci_no_col)
                if isinstance(ogrenci_no_raw, float):
                    ogrenci_no = str(int(ogrenci_no_raw)).strip()
                else:
                    ogrenci_no = str(ogrenci_no_raw).strip() if ogrenci_no_raw else None
                
                # Boş veya geçersiz öğrenci numarası atla
                if not ogrenci_no or ogrenci_no == '0' or ogrenci_no == '':
                    continue
                
                # Ad Soyadı al ve ayır
                ad_soyad = str(ws.cell_value(satir, ad_soyad_col)).strip() if ws.cell_value(satir, ad_soyad_col) else None
                
                if not ad_soyad:
                    hata += 1
                    continue
                
                # Ad Soyadı ayır (son kelime soyad, kalanlar ad)
                parcalar = ad_soyad.split()
                if len(parcalar) >= 2:
                    soyad = parcalar[-1]
                    ad = ' '.join(parcalar[:-1])
                else:
                    ad = ad_soyad
                    soyad = ''
                
                if not ad:
                    hata += 1
                    continue
                
                # Öğrenci var mı kontrol et
                ogrenci = Ogrenci.query.filter_by(ogrenci_no=ogrenci_no).first()
                
                if not ogrenci:
                    # Yeni öğrenci oluştur
                    ogrenci = Ogrenci(
                        ogrenci_no=ogrenci_no,
                        ad=ad,
                        soyad=soyad,
                        email=None,
                        bolum_id=ders.bolum_id
                    )
                    db.session.add(ogrenci)
                    db.session.flush()  # ID'yi al
                
                # Öğrenciyi derse kaydet (zaten kayıtlı değilse)
                if not OgrenciDers.query.filter_by(ogrenci_id=ogrenci.id, ders_id=ders.id).first():
                    ogrenci_ders = OgrenciDers(
                        ogrenci_id=ogrenci.id,
                        ders_id=ders.id
                    )
                    db.session.add(ogrenci_ders)
                    eklenen += 1
        finally:
            pass  # xlrd workbook'u otomatik kapanır
    
    else:
        # .xlsx formatı - openpyxl kullan
        wb = load_workbook(dosya_yolu)
        try:
            ws = wb.active
            
            # İlk satır başlık olabilir
            baslangic = 1
            if ws['A1'].value and str(ws['A1'].value).lower() in ['ogrenci_no', 'öğrenci_no', 'numara', 'no']:
                baslangic = 2
            
            for satir in range(baslangic, ws.max_row + 1):
                ogrenci_no = str(ws[f'A{satir}'].value).strip() if ws[f'A{satir}'].value else None
                ad = str(ws[f'B{satir}'].value).strip() if ws[f'B{satir}'].value else None
                soyad = str(ws[f'C{satir}'].value).strip() if ws[f'C{satir}'].value else None
                email = str(ws[f'D{satir}'].value).strip() if ws[f'D{satir}'].value and ws[f'D{satir}'].value != 'None' else None
                
                if not ogrenci_no or not ad or not soyad:
                    hata += 1
                    continue
                
                # Öğrenci var mı kontrol et
                ogrenci = Ogrenci.query.filter_by(ogrenci_no=ogrenci_no).first()
                
                if not ogrenci:
                    # Yeni öğrenci oluştur
                    ogrenci = Ogrenci(
                        ogrenci_no=ogrenci_no,
                        ad=ad,
                        soyad=soyad,
                        email=email,
                        bolum_id=ders.bolum_id
                    )
                    db.session.add(ogrenci)
                    db.session.flush()  # ID'yi al
                
                # Öğrenciyi derse kaydet (zaten kayıtlı değilse)
                if not OgrenciDers.query.filter_by(ogrenci_id=ogrenci.id, ders_id=ders.id).first():
                    ogrenci_ders = OgrenciDers(
                        ogrenci_id=ogrenci.id,
                        ders_id=ders.id
                    )
                    db.session.add(ogrenci_ders)
                    eklenen += 1
        finally:
            # Excel dosyasını kapat
            wb.close()
    
    db.session.commit()
    
    return {'eklenen': eklenen, 'hata': hata}
