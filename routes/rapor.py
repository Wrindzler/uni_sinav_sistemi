"""
RAPORLAMA ROUTE'LARI
====================
PDF ve Excel rapor oluşturma işlemleri.

Tüm kullanıcılar rapor alabilir.
"""

from flask import Blueprint, render_template, request, send_file, jsonify
from flask_login import login_required, current_user
from datetime import datetime, date
from sqlalchemy.orm import joinedload
from models.sinav import Sinav
from models.ders import Ders
from models.derslik import Derslik
from models.ogretim_uyesi import OgretimUyesi
from models.bolum import Bolum
from models.fakulte import Fakulte
from models.ogrenci_ders import OgrenciDers
from models.ogrenci import Ogrenci
from utils.helpers import tarih_formatla, saat_formatla
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Blueprint oluştur
rapor_bp = Blueprint('rapor', __name__, url_prefix='/rapor')


@rapor_bp.route('/')
@login_required
def index():
    """
    Rapor ana sayfası.
    
    Rapor seçeneklerini gösterir.
    """
    fakulteler = Fakulte.query.all()
    bolumler = Bolum.query.all()
    
    return render_template('rapor/index.html', fakulteler=fakulteler, bolumler=bolumler)


@rapor_bp.route('/pdf')
@login_required
def pdf():
    """
    PDF rapor oluştur ve indir.
    
    Filtreleme parametrelerine göre PDF rapor oluşturur.
    """
    # Filtreleme parametreleri
    fakulte_id = request.args.get('fakulte_id', type=int)
    bolum_id = request.args.get('bolum_id', type=int)
    baslangic_tarihi = request.args.get('baslangic_tarihi')
    bitis_tarihi = request.args.get('bitis_tarihi')
    
    # Sorgu oluştur (sinav.py'deki ile aynı mantık)
    sorgu = Sinav.query.filter_by(durum='planlandi')
    
    # Kullanıcı rolüne göre filtreleme
    if current_user.is_ogrenci():
        # Öğrenci ise: Sadece kayıtlı olduğu derslerin sınavlarını göster
        ogrenci = Ogrenci.query.get(current_user.ogrenci_id)
        if ogrenci:
            ogrenci_ders_kayitlari = OgrenciDers.query.filter_by(ogrenci_id=ogrenci.id).all()
            ders_ids = [od.ders_id for od in ogrenci_ders_kayitlari]
            if ders_ids:
                sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
            else:
                # Öğrencinin kayıtlı dersi yoksa boş liste döndür
                sorgu = sorgu.filter(Sinav.ders_id.in_([]))
        else:
            sorgu = sorgu.filter(Sinav.ders_id.in_([]))
    
    elif current_user.is_hoca():
        # Hoca ise: Sadece verdiği derslerin sınavlarını göster
        ogretim_uyesi = OgretimUyesi.query.get(current_user.ogretim_uyesi_id)
        if ogretim_uyesi:
            ders_ids = [d.id for d in Ders.query.filter_by(ogretim_uyesi_id=ogretim_uyesi.id, aktif=True).all()]
            if ders_ids:
                sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
            else:
                # Hocanın dersi yoksa boş liste döndür
                sorgu = sorgu.filter(Sinav.ders_id.in_([]))
        else:
            sorgu = sorgu.filter(Sinav.ders_id.in_([]))
    
    else:
        # Admin veya bölüm yetkilisi için mevcut filtreleme mantığı
        if bolum_id:
            ders_ids = [d.id for d in Ders.query.filter_by(bolum_id=bolum_id).all()]
            sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
        elif fakulte_id:
            bolum_ids = [b.id for b in Bolum.query.filter_by(fakulte_id=fakulte_id).all()]
            ders_ids = [d.id for d in Ders.query.filter(Ders.bolum_id.in_(bolum_ids)).all()]
            sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
    
    if baslangic_tarihi:
        sorgu = sorgu.filter(Sinav.tarih >= datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date())
    
    if bitis_tarihi:
        sorgu = sorgu.filter(Sinav.tarih <= datetime.strptime(bitis_tarihi, '%Y-%m-%d').date())

    # N+1 query problemini önlemek için joinedload kullan
    sinavlar = sorgu.options(
        joinedload(Sinav.ders).joinedload(Ders.ogretim_uyesi),
        joinedload(Sinav.derslik)
    ).order_by(Sinav.tarih, Sinav.baslangic_saati).all()

    # PDF oluştur
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Türkçe karakter desteği için DejaVu Sans font'u kaydet (gömülü)
    # ReportLab'ın dahili Unicode desteği kullanılacak
    try:
        # Windows için
        pdfmetrics.registerFont(TTFont('Turkish', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Turkish-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        font_name = 'Turkish'
        font_bold = 'Turkish-Bold'
    except:
        # Arial bulunamazsa Helvetica kullan (sınırlı Türkçe desteği)
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
    
    # Stil tanımlamaları
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_bold,
        fontSize=18,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=30,
        alignment=1  # Ortala
    )
    
    # Başlık
    story.append(Paragraph('ÜNİVERSİTE SINAV PROGRAMI', title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Tarih bilgisi
    if baslangic_tarihi and bitis_tarihi:
        tarih_metni = f'Tarih Aralığı: {baslangic_tarihi} - {bitis_tarihi}'
    else:
        tarih_metni = f'Rapor Tarihi: {date.today()}'
    
    story.append(Paragraph(tarih_metni, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Tablo verileri
    table_data = [['Tarih', 'Saat', 'Ders', 'Öğretim Üyesi', 'Derslik', 'Öğrenci Sayısı']]
    
    for sinav in sinavlar:
        table_data.append([
            tarih_formatla(sinav.tarih),
            f'{saat_formatla(sinav.baslangic_saati)} - {saat_formatla(sinav.bitis_saati)}',
            sinav.ders.ad,
            sinav.ders.ogretim_uyesi.tam_ad,
            sinav.derslik.ad,
            str(sinav.ders.ogrenci_sayisi)
        ])
    
    # Tablo oluştur
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    
    story.append(table)
    
    # PDF'i oluştur
    doc.build(story)
    buffer.seek(0)
    
    # Dosya adı
    dosya_adi = f'sinav_programi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=dosya_adi)


@rapor_bp.route('/excel')
@login_required
def excel():
    """
    Excel rapor oluştur ve indir.
    
    Filtreleme parametrelerine göre Excel rapor oluşturur.
    """
    # Filtreleme parametreleri (PDF ile aynı)
    fakulte_id = request.args.get('fakulte_id', type=int)
    bolum_id = request.args.get('bolum_id', type=int)
    baslangic_tarihi = request.args.get('baslangic_tarihi')
    bitis_tarihi = request.args.get('bitis_tarihi')
    
    # Sorgu oluştur
    sorgu = Sinav.query.filter_by(durum='planlandi')
    
    # Kullanıcı rolüne göre filtreleme (PDF ile aynı mantık)
    if current_user.is_ogrenci():
        # Öğrenci ise: Sadece kayıtlı olduğu derslerin sınavlarını göster
        ogrenci = Ogrenci.query.get(current_user.ogrenci_id)
        if ogrenci:
            ogrenci_ders_kayitlari = OgrenciDers.query.filter_by(ogrenci_id=ogrenci.id).all()
            ders_ids = [od.ders_id for od in ogrenci_ders_kayitlari]
            if ders_ids:
                sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
            else:
                # Öğrencinin kayıtlı dersi yoksa boş liste döndür
                sorgu = sorgu.filter(Sinav.ders_id.in_([]))
        else:
            sorgu = sorgu.filter(Sinav.ders_id.in_([]))
    
    elif current_user.is_hoca():
        # Hoca ise: Sadece verdiği derslerin sınavlarını göster
        ogretim_uyesi = OgretimUyesi.query.get(current_user.ogretim_uyesi_id)
        if ogretim_uyesi:
            ders_ids = [d.id for d in Ders.query.filter_by(ogretim_uyesi_id=ogretim_uyesi.id, aktif=True).all()]
            if ders_ids:
                sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
            else:
                # Hocanın dersi yoksa boş liste döndür
                sorgu = sorgu.filter(Sinav.ders_id.in_([]))
        else:
            sorgu = sorgu.filter(Sinav.ders_id.in_([]))
    
    else:
        # Admin veya bölüm yetkilisi için mevcut filtreleme mantığı
        if bolum_id:
            ders_ids = [d.id for d in Ders.query.filter_by(bolum_id=bolum_id).all()]
            sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
        elif fakulte_id:
            bolum_ids = [b.id for b in Bolum.query.filter_by(fakulte_id=fakulte_id).all()]
            ders_ids = [d.id for d in Ders.query.filter(Ders.bolum_id.in_(bolum_ids)).all()]
            sorgu = sorgu.filter(Sinav.ders_id.in_(ders_ids))
    
    if baslangic_tarihi:
        sorgu = sorgu.filter(Sinav.tarih >= datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date())
    
    if bitis_tarihi:
        sorgu = sorgu.filter(Sinav.tarih <= datetime.strptime(bitis_tarihi, '%Y-%m-%d').date())

    # N+1 query problemini önlemek için joinedload kullan
    sinavlar = sorgu.options(
        joinedload(Sinav.ders).joinedload(Ders.ogretim_uyesi),
        joinedload(Sinav.derslik)
    ).order_by(Sinav.tarih, Sinav.baslangic_saati).all()

    # Excel çalışma kitabı oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sınav Programı'
    
    # Başlık satırı
    basliklar = ['Tarih', 'Gün', 'Başlangıç', 'Bitiş', 'Ders', 'Ders Kodu', 
                 'Öğretim Üyesi', 'Derslik', 'Öğrenci Sayısı']
    
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=baslik)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Veri satırları
    gun_isimleri = {
        'Monday': 'Pazartesi',
        'Tuesday': 'Salı',
        'Wednesday': 'Çarşamba',
        'Thursday': 'Perşembe',
        'Friday': 'Cuma',
        'Saturday': 'Cumartesi',
        'Sunday': 'Pazar'
    }
    
    for row, sinav in enumerate(sinavlar, 2):
        gun_adi = sinav.tarih.strftime('%A')
        turkce_gun = gun_isimleri.get(gun_adi, gun_adi)
        
        ws.cell(row=row, column=1, value=sinav.tarih.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=2, value=turkce_gun)
        ws.cell(row=row, column=3, value=sinav.baslangic_saati.strftime('%H:%M'))
        ws.cell(row=row, column=4, value=sinav.bitis_saati.strftime('%H:%M'))
        ws.cell(row=row, column=5, value=sinav.ders.ad)
        ws.cell(row=row, column=6, value=sinav.ders.kod or '')
        ws.cell(row=row, column=7, value=sinav.ders.ogretim_uyesi.tam_ad)
        ws.cell(row=row, column=8, value=sinav.derslik.ad)
        ws.cell(row=row, column=9, value=sinav.ders.ogrenci_sayisi)
    
    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    # Excel dosyasını buffer'a kaydet
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Dosya adı
    dosya_adi = f'sinav_programi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=dosya_adi)

