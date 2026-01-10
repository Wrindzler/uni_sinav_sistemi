"""
YARDIMCI FONKSİYONLAR
=====================
Genel amaçlı yardımcı fonksiyonlar.

Bu modül:
- Tarihleri ve saatleri Türkçe formatında döndürür
- Doğrulama işlemleri yapar
- Veri dönüşümleri gerçekleştirir
- Format işlemleri yapılır

İçeriği:
--------
- tarih_formatla: Tarihi Türkçe format'ta göster
- saat_formatla: Saati HH:MM formatında göster
- gun_adi_tr: Haftanın gün adını Türkçe olarak döndür
- validate_email: E-posta doğrulaması
- validate_csv_file: CSV dosyası doğrulaması
"""

from datetime import date, time
from typing import Optional, List
import re


# ============================================
# TARİH VE SAAT FORMATLAMA FONKSİYONLARI
# ============================================

def tarih_formatla(tarih: Optional[date]) -> str:
    """
    Bir tarih nesnesini Türkçe formatında string'e çevir.
    
    Tarih nesnesi None ise boş string döner.
    
    Args:
        tarih (date): Formatlanacak tarih nesnesi
        
    Returns:
        str: Formatlanmış tarih (örn: "15 Ocak 2024")
        
    Örnekler:
        >>> from datetime import date
        >>> tarih = date(2024, 1, 15)
        >>> tarih_formatla(tarih)
        '15 Ocak 2024'
        
        >>> tarih_formatla(None)
        ''
    """
    if not tarih:
        return ''
    
    # Ayların Türkçe isimleri (1-indexed, 0. index boş)
    aylar = [
        '',           # 0: kullanılmaz
        'Ocak',       # 1: January
        'Şubat',      # 2: February
        'Mart',       # 3: March
        'Nisan',      # 4: April
        'Mayıs',      # 5: May
        'Haziran',    # 6: June
        'Temmuz',     # 7: July
        'Ağustos',    # 8: August
        'Eylül',      # 9: September
        'Ekim',       # 10: October
        'Kasım',      # 11: November
        'Aralık'      # 12: December
    ]
    
    # Formatı: gün + ayın adı + yıl
    return f'{tarih.day} {aylar[tarih.month]} {tarih.year}'


def saat_formatla(saat: Optional[time]) -> str:
    """
    Bir saat nesnesini HH:MM formatında string'e çevir.
    
    Saat nesnesi None ise boş string döner.
    
    Args:
        saat (time): Formatlanacak saat nesnesi
        
    Returns:
        str: Formatlanmış saat (örn: "14:30")
        
    Örnekler:
        >>> from datetime import time
        >>> saat = time(14, 30)
        >>> saat_formatla(saat)
        '14:30'
        
        >>> saat = time(9, 5)
        >>> saat_formatla(saat)
        '09:05'
        
        >>> saat_formatla(None)
        ''
    """
    if not saat:
        return ''
    
    # strftime: %H (saat, 00-23), %M (dakika, 00-59)
    return saat.strftime('%H:%M')


def gun_adi_tr(tarih: Optional[date]) -> str:
    """
    Bir tarihin haftanın hangi günü olduğunu Türkçe olarak döndür.
    
    Python'ın weekday() metodu:
    - 0: Pazartesi
    - 1: Salı
    - 2: Çarşamba
    - 3: Perşembe
    - 4: Cuma
    - 5: Cumartesi
    - 6: Pazar
    
    Args:
        tarih (date): Gün adını bulacak tarih nesnesi
        
    Returns:
        str: Gün adı Türkçe (örn: "Pazartesi")
        
    Örnekler:
        >>> from datetime import date
        >>> tarih = date(2024, 1, 15)  # Pazartesi
        >>> gun_adi_tr(tarih)
        'Pazartesi'
        
        >>> tarih = date(2024, 1, 20)  # Cumartesi
        >>> gun_adi_tr(tarih)
        'Cumartesi'
        
        >>> gun_adi_tr(None)
        ''
    """
    if not tarih:
        return ''
    
    # Gün isimleri (Pazartesi = 0, Pazar = 6)
    gunler = [
        'Pazartesi',    # 0
        'Salı',         # 1
        'Çarşamba',     # 2
        'Perşembe',     # 3
        'Cuma',         # 4
        'Cumartesi',    # 5
        'Pazar'         # 6
    ]
    
    # weekday() metoduyla gün indeksini bul ve adını döndür
    return gunler[tarih.weekday()]


# ============================================
# DOĞRULAMA FONKSİYONLARI
# ============================================

def validate_email(email: str) -> bool:
    """
    E-posta adresinin geçerli olup olmadığını kontrol et.
    
    Basit bir regex deseni kullanarak e-posta formatını doğrular.
    Gerçek doğrulama için daha karmaşık pattern'ler kullanılabilir.
    
    Args:
        email (str): Kontrol edilecek e-posta adresi
        
    Returns:
        bool: Geçerli ise True, değilse False
        
    Örnekler:
        >>> validate_email('user@example.com')
        True
        
        >>> validate_email('invalid.email')
        False
        
        >>> validate_email('test@universite.edu.tr')
        True
    """
    # Basit e-posta regex deseni
    # Deseni açıkla: 
    # ^[a-zA-Z0-9._%+-]+    : Başlangıç, izin verilen karakterler
    # @                      : At işareti zorunlu
    # [a-zA-Z0-9.-]+         : Domain adı
    # \.[a-zA-Z]{2,}$        : Nokta ve en az 2 karakter TLD
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    
    return re.match(pattern, email) is not None


def validate_ogrenci_no(ogrenci_no: str) -> bool:
    """
    Öğrenci numarasının formatını kontrol et.
    
    Tipik format: 7 basamaklı sayı (örn: 2010001)
    Veya: Başında yıl ile başlayan (örn: 201-0001, 20-10001)
    
    Args:
        ogrenci_no (str): Kontrol edilecek öğrenci numarası
        
    Returns:
        bool: Geçerli ise True, değilse False
        
    Örnekler:
        >>> validate_ogrenci_no('2010001')
        True
        
        >>> validate_ogrenci_no('201-0001')
        True
        
        >>> validate_ogrenci_no('invalid')
        False
    """
    # Formatı: 7 basamak veya 3-4 basamak formatında
    pattern = r'^[\d]{7}$|^[\d]{3}-[\d]{4}$|^[\d]{2}-[\d]{5}$'
    
    return re.match(pattern, ogrenci_no) is not None


def validate_csv_file(filename: str) -> bool:
    """
    Dosya adının CSV formatında olup olmadığını kontrol et.
    
    Args:
        filename (str): Kontrol edilecek dosya adı
        
    Returns:
        bool: CSV dosyası ise True
        
    Örnekler:
        >>> validate_csv_file('ogrenciler.csv')
        True
        
        >>> validate_csv_file('veriler.xlsx')
        False
    """
    return filename.lower().endswith('.csv')


# ============================================
# VERİ DÖNÜŞÜMÜ FONKSİYONLARI
# ============================================

def saniye_to_dakika_saat(saniye: int) -> str:
    """
    Saniyeyi saat ve dakikaya çevir.
    
    Args:
        saniye (int): Dönüştürülecek saniye
        
    Returns:
        str: Formatlanmış saat ve dakika (örn: "1 saat 30 dakika")
        
    Örnekler:
        >>> saniye_to_dakika_saat(3600)  # 1 saat
        '1 saat 0 dakika'
        
        >>> saniye_to_dakika_saat(5400)  # 1.5 saat
        '1 saat 30 dakika'
    """
    saat = saniye // 3600
    dakika = (saniye % 3600) // 60
    
    return f'{saat} saat {dakika} dakika'


def dakika_to_saat_dakika(dakika: int) -> tuple:
    """
    Dakikayı saat ve dakikaya çevir.
    
    Args:
        dakika (int): Dönüştürülecek dakika
        
    Returns:
        tuple: (saat, kalan_dakika)
        
    Örnekler:
        >>> dakika_to_saat_dakika(90)
        (1, 30)
        
        >>> dakika_to_saat_dakika(45)
        (0, 45)
    """
    saat = dakika // 60
    kalan_dakika = dakika % 60
    
    return (saat, kalan_dakika)


# ============================================
# LİSTE İŞLEMİ FONKSİYONLARI
# ============================================

def ozellikler_listesi_olustur(nesne, gizli_ozellikler: List[str] = None) -> dict:
    """
    Bir nesnenin tüm özelliklerini dictionary'ye çevir.
    
    Bazı özellikler gizli tutulabilir.
    
    Args:
        nesne: Dönüştürülecek nesne (model instance)
        gizli_ozellikler (list): Gösterilmeyecek özellik adları
        
    Returns:
        dict: Nesnenin özellikleri
        
    Örnek:
        >>> from models.ogrenci import Ogrenci
        >>> ogrenci = Ogrenci(ogrenci_no='2010001', ad='Ahmet', soyad='Yılmaz')
        >>> ozellikler = ozellikler_listesi_olustur(
        ...     ogrenci, 
        ...     gizli_ozellikler=['id', 'dersler']
        ... )
    """
    if gizli_ozellikler is None:
        gizli_ozellikler = []
    
    sonuc = {}
    for key, value in nesne.__dict__.items():
        # _ ile başlayanlar (private) ve gizli olanları atla
        if not key.startswith('_') and key not in gizli_ozellikler:
            sonuc[key] = value
    
    return sonuc


# ============================================
# BÖLÜM YETKİLİSİ YARDIMCI FONKSİYONLARI
# ============================================

def get_bolum_yetkilisi_bolum_id(current_user):
    """
    Bölüm yetkilisinin bölüm ID'sini döndür.
    
    Args:
        current_user: Mevcut kullanıcı
        
    Returns:
        int or None: Bölüm ID'si veya None
    """
    from models.bolum_yetkilisi import BolumYetkilisi
    
    if not current_user.is_bolum_yetkilisi():
        return None
    
    bolum_yetkilisi = BolumYetkilisi.query.get(current_user.bolum_yetkilisi_id)
    if bolum_yetkilisi:
        return bolum_yetkilisi.bolum_id
    return None


# ============================================
# DOSYA YÜKLEME YARDIMCI FONKSİYONLARI
# ============================================

def parse_csv_file(dosya_yolu):
    """
    CSV dosyasını parse et ve satırları döndür.
    
    Args:
        dosya_yolu: CSV dosyasının yolu
        
    Returns:
        list: Satırlar listesi (başlık satırı hariç)
    """
    import csv
    
    with open(dosya_yolu, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        satirlar = list(reader)
        
        # İlk satır başlık olabilir, kontrol et
        baslangic = 0
        if satirlar and (satirlar[0][0].lower() in ['ogrenci_no', 'öğrenci_no', 'numara', 'no']):
            baslangic = 1
        
        return satirlar[baslangic:]


def parse_excel_file(dosya_yolu):
    """
    Excel dosyasını parse et ve satırları döndür.
    
    Args:
        dosya_yolu: Excel dosyasının yolu
        
    Returns:
        tuple: (başlangıç_satırı, worksheet, workbook)
    """
    if dosya_yolu.endswith('.xls'):
        # Eski .xls formatı - xlrd kullan
        import xlrd
        wb = xlrd.open_workbook(dosya_yolu)
        ws = wb.sheet_by_index(0)
        
        # İlk satır başlık olabilir
        baslangic = 0
        if ws.nrows > 0 and ws.cell_value(0, 0) and str(ws.cell_value(0, 0)).lower() in ['ogrenci_no', 'öğrenci_no', 'numara', 'no']:
            baslangic = 1
        
        return baslangic, ws, wb
    else:
        # .xlsx formatı - openpyxl kullan
        from openpyxl import load_workbook
        
        wb = load_workbook(dosya_yolu)
        ws = wb.active
        
        # İlk satır başlık olabilir
        baslangic = 1
        if ws['A1'].value and str(ws['A1'].value).lower() in ['ogrenci_no', 'öğrenci_no', 'numara', 'no']:
            baslangic = 2
        
        return baslangic, ws, wb
