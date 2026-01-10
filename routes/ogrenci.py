"""
ÖĞRENCİ YÖNETİMİ ROUTE'LARI
============================
Öğrenci ekleme, silme, güncelleme ve toplu yükleme işlemleri.

Admin ve bölüm yetkilileri öğrenci yönetimi yapabilir.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from models.database import db
from models.ogrenci import Ogrenci
from models.ogrenci_ders import OgrenciDers
from models.ders import Ders
from werkzeug.utils import secure_filename
from utils.decorators import bolum_yetkilisi_required
from utils.helpers import get_bolum_yetkilisi_bolum_id
import os
import csv
from openpyxl import load_workbook
from config import Config

# Blueprint oluştur
ogrenci_bp = Blueprint('ogrenci', __name__, url_prefix='/ogrenci')

# İzin verilen dosya uzantıları
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}


def allowed_file(filename):
    """
    Dosya uzantısının izin verilenler arasında olup olmadığını kontrol et.
    
    Args:
        filename: Dosya adı
        
    Returns:
        bool: İzin verilen uzantı ise True
    """
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@ogrenci_bp.route('/')
@login_required
@bolum_yetkilisi_required
def liste():
    """
    Öğrenci listesi sayfası.

    Tüm öğrencileri veya bölüme ait öğrencileri listeler.
    """
    # Admin ise tüm öğrenciler, bölüm yetkilisi ise sadece kendi bölümünün öğrencileri
    if current_user.is_admin():
        ogrenciler = Ogrenci.query.all()
    else:
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if bolum_id:
            ogrenciler = Ogrenci.query.filter_by(bolum_id=bolum_id).all()
        else:
            ogrenciler = []

    return render_template('ogrenci/liste.html', ogrenciler=ogrenciler)


@ogrenci_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
@bolum_yetkilisi_required
def ekle():
    """
    Yeni öğrenci ekleme sayfası.

    Artık kullanılmıyor - auth.register sayfasına yönlendirir.
    """
    # Yeni kullanıcı kayıt sayfasına yönlendir
    flash('Öğrenci eklemek için kullanıcı kayıt sayfasını kullanın.', 'info')
    return redirect(url_for('auth.register'))


@ogrenci_bp.route('/yukle', methods=['GET', 'POST'])
@login_required
@bolum_yetkilisi_required
def yukle():
    """
    Excel/CSV dosyasından toplu öğrenci yükleme sayfası.
    
    GET: Yükleme formunu gösterir
    POST: Dosyayı işler ve öğrencileri ekler
    """
    if request.method == 'POST':
        # Dosya kontrolü
        if 'dosya' not in request.files:
            flash('Dosya seçilmedi!', 'error')
            return redirect(url_for('ogrenci.yukle'))
        
        dosya = request.files['dosya']
        
        if dosya.filename == '':
            flash('Dosya seçilmedi!', 'error')
            return redirect(url_for('ogrenci.yukle'))
        
        if not allowed_file(dosya.filename):
            flash('Geçersiz dosya formatı! Sadece CSV, XLSX veya XLS dosyaları yüklenebilir.', 'error')
            return redirect(url_for('ogrenci.yukle'))
        
        # Uploads klasörünü oluştur (yoksa)
        uploads_dir = Config.UPLOAD_FOLDER
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)
        
        # Dosyayı kaydet
        filename = secure_filename(dosya.filename)
        dosya_yolu = os.path.join(uploads_dir, filename)
        dosya.save(dosya_yolu)
        
        # Dosyayı işle
        try:
            sonuc = _dosyadan_ogrenci_yukle(dosya_yolu, filename)
            
            # Dosyayı sil (hata olsa bile devam et)
            try:
                os.remove(dosya_yolu)
            except PermissionError:
                # Windows'ta dosya hala açıksa silme, sorun değil
                pass
            
            flash(f'{sonuc["eklenen"]} öğrenci eklendi, {sonuc["hata"]} hata oluştu.', 
                  'success' if sonuc['hata'] == 0 else 'warning')
            return redirect(url_for('ogrenci.liste'))
        except Exception as e:
            # Dosyayı sil (hata olsa bile devam et)
            try:
                if os.path.exists(dosya_yolu):
                    os.remove(dosya_yolu)
            except PermissionError:
                pass
            flash(f'Dosya işlenirken hata oluştu: {str(e)}', 'error')
            return redirect(url_for('ogrenci.yukle'))
    
    return render_template('ogrenci/yukle.html')


@ogrenci_bp.route('/<int:ogrenci_id>/derslere_kaydet', methods=['GET', 'POST'])
@login_required
@bolum_yetkilisi_required
def derslere_kaydet(ogrenci_id):
    """
    Öğrenciyi derslere kaydetme sayfası.
    
    Args:
        ogrenci_id: Kaydedilecek öğrenci ID'si
    """
    ogrenci = Ogrenci.query.get_or_404(ogrenci_id)
    
    if request.method == 'POST':
        # Seçilen dersleri al (tiklenen kutucuklar)
        secilen_ders_ids = [int(d) for d in request.form.getlist('ders_ids')]
        
        # Mevcut kayıtlı dersleri al
        mevcut_kayitlar = OgrenciDers.query.filter_by(ogrenci_id=ogrenci_id).all()
        mevcut_ders_ids = [k.ders_id for k in mevcut_kayitlar]
        
        # Eklenecek dersler: Seçili ama kayıtlı olmayanlar
        eklenecek_dersler = [d_id for d_id in secilen_ders_ids if d_id not in mevcut_ders_ids]
        
        # Silinecek dersler: Kayıtlı ama seçili olmayanlar
        silinecek_dersler = [d_id for d_id in mevcut_ders_ids if d_id not in secilen_ders_ids]
        
        # Yeni kayıtları ekle
        eklenen = 0
        for ders_id in eklenecek_dersler:
            ogrenci_ders = OgrenciDers(
                ogrenci_id=ogrenci_id,
                ders_id=ders_id
            )
            db.session.add(ogrenci_ders)
            eklenen += 1
        
        # Kayıtları sil
        silinen = 0
        for ders_id in silinecek_dersler:
            ogrenci_ders = OgrenciDers.query.filter_by(
                ogrenci_id=ogrenci_id,
                ders_id=ders_id
            ).first()
            if ogrenci_ders:
                db.session.delete(ogrenci_ders)
                silinen += 1
        
        db.session.commit()
        
        # Tüm etkilenen derslerin öğrenci sayısını güncelle
        etkilenen_dersler = set(eklenecek_dersler + silinecek_dersler)
        for ders_id in etkilenen_dersler:
            ders = Ders.query.get(ders_id)
            if ders:
                ders.ogrenci_sayisi = OgrenciDers.query.filter_by(ders_id=ders_id).count()
        
        db.session.commit()
        
        # Mesaj göster
        mesajlar = []
        if eklenen > 0:
            mesajlar.append(f'{eklenen} derse kaydedildi')
        if silinen > 0:
            mesajlar.append(f'{silinen} dersten kayıt silindi')
        
        if mesajlar:
            flash(', '.join(mesajlar) + '!', 'success')
        else:
            flash('Değişiklik yapılmadı.', 'info')
        
        return redirect(url_for('ogrenci.derslere_kaydet', ogrenci_id=ogrenci_id))
    
    # Mevcut dersleri al
    if current_user.is_admin():
        dersler = Ders.query.filter_by(aktif=True).all()
    else:
        bolum_id = get_bolum_yetkilisi_bolum_id(current_user)
        if bolum_id:
            dersler = Ders.query.filter_by(bolum_id=bolum_id, aktif=True).all()
        else:
            dersler = []
    
    # Öğrencinin kayıtlı olduğu dersler (detaylı bilgi ile)
    kayitli_dersler = OgrenciDers.query.filter_by(ogrenci_id=ogrenci_id).all()
    kayitli_ders_ids = [od.ders_id for od in kayitli_dersler]
    
    return render_template('ogrenci/derslere_kaydet.html', 
                         ogrenci=ogrenci, 
                         dersler=dersler,
                         kayitli_dersler=kayitli_dersler,
                         kayitli_ders_ids=kayitli_ders_ids)


@ogrenci_bp.route('/<int:ogrenci_id>/ders_kayit_sil/<int:ders_id>', methods=['POST'])
@login_required
@bolum_yetkilisi_required
def ders_kayit_sil(ogrenci_id, ders_id):
    """
    Öğrencinin bir dersten kaydını sil.
    
    Args:
        ogrenci_id: Öğrenci ID'si
        ders_id: Ders ID'si
    """
    try:
        ogrenci = Ogrenci.query.get_or_404(ogrenci_id)
        ders = Ders.query.get_or_404(ders_id)
        
        # Kayıt var mı kontrol et
        ogrenci_ders = OgrenciDers.query.filter_by(
            ogrenci_id=ogrenci_id,
            ders_id=ders_id
        ).first()
        
        if not ogrenci_ders:
            flash('Bu öğrenci bu derse kayıtlı değil!', 'warning')
            return redirect(url_for('ogrenci.derslere_kaydet', ogrenci_id=ogrenci_id))
        
        # Kaydı sil
        db.session.delete(ogrenci_ders)
        db.session.commit()
        
        # Dersin öğrenci sayısını güncelle
        ders.ogrenci_sayisi = OgrenciDers.query.filter_by(ders_id=ders_id).count()
        db.session.commit()
        
        flash(f'"{ders.ad}" dersinden kayıt silindi!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Kayıt silinirken hata oluştu: {str(e)}', 'danger')
    
    return redirect(url_for('ogrenci.derslere_kaydet', ogrenci_id=ogrenci_id))


def _dosyadan_ogrenci_yukle(dosya_yolu, filename):
    """
    Excel veya CSV dosyasından öğrencileri yükle.
    
    Dosya formatı:
    - CSV: ogrenci_no,ad,soyad,email (başlık satırı opsiyonel)
    - Excel: A=ogrenci_no, B=ad, C=soyad, D=email (ilk satır başlık olabilir)
    
    Args:
        dosya_yolu: Dosyanın tam yolu
        filename: Dosya adı
        
    Returns:
        dict: {'eklenen': int, 'hata': int}
    """
    eklenen = 0
    hata = 0
    
    # Dosya uzantısına göre işle
    if filename.endswith('.csv'):
        # CSV dosyasını oku
        with open(dosya_yolu, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            satirlar = list(reader)
            
            # İlk satır başlık olabilir, kontrol et
            baslangic = 0
            if satirlar and (satirlar[0][0].lower() in ['ogrenci_no', 'öğrenci_no', 'numara', 'no']):
                baslangic = 1
            
            for satir in satirlar[baslangic:]:
                if len(satir) < 3:  # En az ogrenci_no, ad, soyad olmalı
                    hata += 1
                    continue
                
                ogrenci_no = satir[0].strip()
                ad = satir[1].strip()
                soyad = satir[2].strip()
                email = satir[3].strip() if len(satir) > 3 else None
                
                if not ogrenci_no or not ad or not soyad:
                    hata += 1
                    continue
                
                # Öğrenci zaten var mı?
                if Ogrenci.query.filter_by(ogrenci_no=ogrenci_no).first():
                    hata += 1
                    continue
                
                # Yeni öğrenci oluştur
                ogrenci = Ogrenci(
                    ogrenci_no=ogrenci_no,
                    ad=ad,
                    soyad=soyad,
                    email=email
                )
                db.session.add(ogrenci)
                eklenen += 1
    
    elif filename.endswith('.xls'):
        # Eski .xls formatı - xlrd kullan
        import xlrd
        wb = xlrd.open_workbook(dosya_yolu)
        try:
            ws = wb.sheet_by_index(0)
            
            # Başlık satırını ve sütun indekslerini bul
            baslik_satiri = -1
            ogrenci_no_col = -1
            ad_soyad_col = -1
            
            for i in range(min(20, ws.nrows)):  # İlk 20 satırda başlık ara
                for j in range(ws.ncols):
                    cell_val = str(ws.cell_value(i, j)).lower().strip()
                    # Encoding sorunları için farklı varyasyonları dene
                    if 'renci no' in cell_val or cell_val == '#':
                        if cell_val == '#':
                            # # sütunu bulundu, öğrenci no genelde 4. sütunda
                            baslik_satiri = i
                        else:
                            baslik_satiri = i
                            ogrenci_no_col = j
                    if 'soyad' in cell_val:
                        ad_soyad_col = j
                if baslik_satiri >= 0 and ogrenci_no_col < 0:
                    # # sütunu bulundu ama öğrenci no sütunu bulunamadı, varsayılan kullan
                    ogrenci_no_col = 4  # Genelde E sütunu
                    ad_soyad_col = 5    # Genelde F sütunu
                if baslik_satiri >= 0:
                    break
            
            # Başlık bulunamazsa varsayılan değerler kullan
            if baslik_satiri < 0:
                baslik_satiri = 4  # Genelde 5. satır başlık
                ogrenci_no_col = 4
                ad_soyad_col = 5
            
            baslangic = baslik_satiri + 1
            
            for satir in range(baslangic, ws.nrows):
                # Öğrenci numarasını al
                ogrenci_no_raw = ws.cell_value(satir, ogrenci_no_col)
                if isinstance(ogrenci_no_raw, float):
                    ogrenci_no = str(int(ogrenci_no_raw)).strip()
                else:
                    ogrenci_no = str(ogrenci_no_raw).strip() if ogrenci_no_raw else None
                
                # Boş veya geçersiz öğrenci numarası atla
                if not ogrenci_no or ogrenci_no == '0' or ogrenci_no == '':
                    continue
                
                # Ad Soyadı al ve ayır
                ad_soyad = str(ws.cell_value(satir, ad_soyad_col)).strip() if ws.cell_value(satir, ad_soyad_col) else None
                
                if not ad_soyad:
                    hata += 1
                    continue
                
                # Ad Soyadı ayır (son kelime soyad, kalanlar ad)
                parcalar = ad_soyad.split()
                if len(parcalar) >= 2:
                    soyad = parcalar[-1]
                    ad = ' '.join(parcalar[:-1])
                else:
                    ad = ad_soyad
                    soyad = ''
                
                if not ad:
                    hata += 1
                    continue
                
                # Öğrenci zaten var mı?
                if Ogrenci.query.filter_by(ogrenci_no=ogrenci_no).first():
                    hata += 1
                    continue
                
                # Yeni öğrenci oluştur
                ogrenci = Ogrenci(
                    ogrenci_no=ogrenci_no,
                    ad=ad,
                    soyad=soyad,
                    email=None
                )
                db.session.add(ogrenci)
                eklenen += 1
        finally:
            pass  # xlrd workbook'u otomatik kapanır
    
    else:
        # .xlsx formatı - openpyxl kullan
        wb = load_workbook(dosya_yolu)
        try:
            ws = wb.active
            
            # İlk satır başlık olabilir
            baslangic = 1
            if ws['A1'].value and str(ws['A1'].value).lower() in ['ogrenci_no', 'öğrenci_no', 'numara', 'no']:
                baslangic = 2
            
            for satir in range(baslangic, ws.max_row + 1):
                ogrenci_no = str(ws[f'A{satir}'].value).strip() if ws[f'A{satir}'].value else None
                ad = str(ws[f'B{satir}'].value).strip() if ws[f'B{satir}'].value else None
                soyad = str(ws[f'C{satir}'].value).strip() if ws[f'C{satir}'].value else None
                email = str(ws[f'D{satir}'].value).strip() if ws[f'D{satir}'].value and ws[f'D{satir}'].value else None
                
                if not ogrenci_no or not ad or not soyad:
                    hata += 1
                    continue
                
                # Öğrenci zaten var mı?
                if Ogrenci.query.filter_by(ogrenci_no=ogrenci_no).first():
                    hata += 1
                    continue
                
                # Yeni öğrenci oluştur
                ogrenci = Ogrenci(
                    ogrenci_no=ogrenci_no,
                    ad=ad,
                    soyad=soyad,
                    email=email
                )
                db.session.add(ogrenci)
                eklenen += 1
        finally:
            # Excel dosyasını kapat
            wb.close()
    
    db.session.commit()

    return {'eklenen': eklenen, 'hata': hata}


@ogrenci_bp.route('/sil/<int:ogrenci_id>', methods=['POST'])
@login_required
@bolum_yetkilisi_required
def ogrenci_sil(ogrenci_id):
    """
    Öğrenci silme işlemi.

    Öğrenciyi ve bağlı kullanıcı hesabını siler.
    """
    try:
        from models.user import User

        ogrenci = Ogrenci.query.get(ogrenci_id)
        if not ogrenci:
            flash('Öğrenci bulunamadı!', 'danger')
            return redirect(url_for('ogrenci.liste'))

        ogrenci_ad = ogrenci.tam_ad

        # İlişkili kullanıcı hesabı varsa onu da sil
        user = User.query.filter_by(ogrenci_id=ogrenci_id).first()
        if user:
            db.session.delete(user)

        # Öğrenciyi sil
        db.session.delete(ogrenci)
        db.session.commit()

        flash(f'"{ogrenci_ad}" öğrencisi başarıyla silindi!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Öğrenci silinirken hata oluştu: {str(e)}', 'danger')

    return redirect(url_for('ogrenci.liste'))

